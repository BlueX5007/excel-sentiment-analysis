import openpyxl
import nltk
from nltk.sentiment import SentimentIntensityAnalyzer

nltk.download('vader_lexicon')

fileName = input("Enter the excel file name without the extension.")
sheetName = input("Enter your sheet name.")
columnNumber = int(input("Enter the column number which will be reserved for review analysis."))
warning = input("By pressing Enter to continue, you hereby acknowledge that you are responsible for your files and that you have read how this code works and modifies your Excel file. Create a backup just for safety.")

wb = openpyxl.load_workbook(f'{fileName}.xlsx')
sheet = wb[sheetName]  # Replace with your sheet name

sia = SentimentIntensityAnalyzer()

for row in range(2, sheet.max_row + 1):  # Assuming row 1 contains headers
    review = sheet.cell(row=row, column=columnNumber).value
    polarity_scores = sia.polarity_scores(review)
    sentiment = 'neutral'
    if polarity_scores['compound'] > 0.05:
        sentiment = 'positive'
    elif polarity_scores['compound'] < -0.05:
        sentiment = 'negative'
    sheet.cell(row=row, column=columnNumber+1, value=sentiment)  # Write sentiment to column 3

wb.save(f'{fileName}.xlsx')